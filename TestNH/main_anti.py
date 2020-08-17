import re,os,sys
import log
import positions
from openpyxl.workbook import Workbook
from openpyxl.styles import Style,alignment,fonts

import win32com.client
from string import ascii_uppercase
from pandas import Series

xlLegendPositionBottom        =-4107
xlLine                        =4          # from enum XlChartType

from wmi import WMI

TRADE_LQ = False

def read_records(trade_rec_file):
	file_name = trade_rec_file
	map_data = []
	index = 1
	f = open(file_name, "rt")
	if not f: 
		err = 'cant open file [%s]' %  file_name
		raise Exception(err)
		return
		
	line = f.readline()
	while line:
		index += 1
		try:	
			ma = re.match("(\d+)\s+(\w+)\s+(\w+)\s+(\S+)\s+(\S+)\s+(\d+)\s+(\d+[\.\d+]*)\s+(\d+[\.\d+]*)\s+(\d+)\s+(\d+)$", line)
			if ma:
				arrs = ma.groups(0)
				tradeDate = int(arrs[0])
				accountID = arrs[1]
				contract = arrs[2]
				trade_dir = arrs[3].decode('utf8')
				trade_type = arrs[4].decode('utf8')
				trade_num = int(arrs[5])
				price = float(arrs[6])
				commission = float(arrs[7])
				orderID = int(arrs[8])
				tradeTime = int(arrs[9])

				map_data.append([accountID, contract, trade_dir, trade_type, price, trade_num, commission, orderID, tradeDate, tradeTime])
			else:
				print("line %d not match!" % (index))

		except Exception as e:
			print(e)

		line = f.readline()
	f.close()

	return map_data

def get_price(datas, datetime):
	if datetime in datas[0]:
		return datas[0][datetime][8]
	else:
		return 0

def get_pre_price(datas, datetime):
	for i in range(0, len(datas[1])):
		if datetime in datas[1][0]:
			return i - 1
	return 0

def get_next_price(datas, datetime):
	for i in range(0, len(datas[1])):
		if datetime in datas[1][0]:
			return i + 1
	return 0

def parse_records(pos_manager, records):
	infos = {
		"contracts" : {},
		"day_profit" : {},
		"day_accountID" : {},
		"day_contracts" : {},
		"day_times" : {}
	}
	contracts = infos["contracts"]
	all_profits = infos["day_profit"]
	day_accountID = infos["day_accountID"]
	day_contracts = infos["day_contracts"]
	day_times = infos["day_times"]

	def func_query_close_price(contract, date):
		return 1

	profit_file = {}
	close_profits = {}

	# 初始化当前日期
	set_accountID = []
	set_contracts = []
	trade_times = 0
	current_date = records[0][8]
	index = 1
	for accountID, contract, dir, type, price, trade_num, commission, orderID, tradeDate, tradeTime in records:
		index += 1
		datetime = tradeDate % 1000000 * 1000000 + tradeTime

		if(not profit_file.has_key(contract)):
			profit_file[contract] = open("profit/%s_profit.db" % contract, "w")

		if current_date != tradeDate: #日期变化，计算日利润
			day_profit = pos_manager.get_position_profit(int(datetime/100), func_query_close_price)
			all_profits[current_date] = day_profit

			day_accountID[current_date] = set_accountID
			set_accountID = []

			day_contracts[current_date] = set_contracts
			set_contracts = []

			day_times[current_date] = trade_times
			trade_times = 0

			current_date = tradeDate

		trade_times = trade_times + 1

		if contract not in set_contracts:
			set_contracts.append(contract)

		if accountID not in set_accountID:
			set_accountID.append(accountID)

		profit = 0
		trade_dir = positions.DIR_BUY
		dir_reverse = positions.DIR_SELL
		if dir.find(u'卖', 0) >= 0:
			trade_dir = positions.DIR_SELL
			dir_reverse = positions.DIR_BUY

		if type.find(u'开', 0) >= 0:
			pos_manager.open(datetime, contract, trade_dir, trade_num, price, commission)
		else:
			profit = pos_manager.close(datetime, contract, trade_dir, trade_num, price, commission, False)			

		if not contracts.has_key(contract): contracts[contract] = 0
		if(close_profits.has_key(contract)):
			close_profits[contract] += profit
		else:
			close_profits[contract] = profit

		profit_file[contract].write("%d\t%d\t%.2f\t%.2f\t%.2f\t%.2f\t%d\n" % (tradeDate%1000000*10000+tradeTime/100, trade_dir, price, profit, close_profits[contract], commission, trade_num))

	# 记录最后一天日利润
	day_profit = pos_manager.get_position_profit(int(datetime/100), func_query_close_price)
	all_profits[current_date] = day_profit

	day_accountID[current_date] = set_accountID
	day_contracts[current_date] = set_contracts
	day_times[current_date] = trade_times

	for k,v in contracts.items():
		contracts[k] = (pos_manager.get_profit(k), pos_manager.get_commission(k))

	# 关闭所有文件
	[v.close() for k,v in profit_file.items()]

	return infos

def make_profit_dir():
	base_path = os.getcwd()
	if not os.path.exists(base_path + "/profit"): os.makedirs(base_path+"/profit") 

def calc_day_profit(contracts, day_contract_profits, day_contracts, day_accountID, day_times):
	contracts.sort()

	# 保存打印
	title = "日期"
	for c in contracts:
		title += "\t" + c

	log.WriteLog("net_profit", title)

	days = day_contract_profits.keys()
	days.sort()
	for day in days:
		line = "%s" % day
		for c in contracts:
			if day_contract_profits[day].has_key(c):
				line += "\t%.2f" % day_contract_profits[day][c]
			else:
				line += "\t0"
		log.WriteLog("net_profit", line)

	title = "日期\t合约"

	log.WriteLog("day_contracts", title)

	days = day_contracts.keys()
	days.sort()
	for day in days:
		line = "%s" % day
		set_contracts = day_contracts[day]
		size = len(set_contracts)
		for i in xrange(0,size):
			if 0 == i:
				line += "\t%s" % set_contracts[i]
			else:
				line += ";%s" % set_contracts[i]
		log.WriteLog("day_contracts", line)

	title = "日期\t账号ID"

	log.WriteLog("day_accountID", title)

	days = day_accountID.keys()
	days.sort()
	for day in days:
		line = "%s" % day
		set_accountID = day_accountID[day]
		size = len(set_accountID)
		for i in xrange(0,size):
			if 0 == i:
				line += "\t%s" % set_accountID[i]
			else:
				line += ";%s" % set_accountID[i]
		log.WriteLog("day_accountID", line)

	title = "日期\t交易次数"

	log.WriteLog("day_times", title)

	days = day_times.keys()
	days.sort()
	for day in days:
		line = "%s\t%d" % (day, day_times[day])
		log.WriteLog("day_times", line)

def do_excel(contracts, day_contract_profits, day_contracts, day_accountID, day_times):
	excel_file = u"log/test.xlsx"

	report_wb = Workbook()

	#在第一个sheet记录基本情况
	report_ws_base = report_wb.worksheets[0]
	report_ws_base.title = u"基本情况"

	#基本情况
	days = day_accountID.keys()
	days.sort()
	days_num = len(days)

	index = 1
	report_ws_base.cell(row = index, column = 1).value = "测试机制：根据交易记录得到每日净值"
	index += 1
	report_ws_base.cell(row = index, column = 1).value = "测试数据时间范围：%u-%u-%u—%u-%u-%u" % (days[0]/10000,days[0]/100%100,days[0]%100,
		days[days_num-1]/10000,days[days_num-1]/100%100,days[days_num-1]%100)

	#记录每日操作账号
	index += 2
	report_ws_base.cell(row = index, column = 1).value = "操作账号情况："
	index += 1
	report_ws_base.cell(row = index, column = 1).value = "日期"
	report_ws_base.cell(row = index, column = 2).value = "账号"
	index += 1
	for day in days:
		new_style = Style(alignment=alignment.Alignment(horizontal='left'),font=fonts.Font(bold=False))

		report_ws_base.cell(row = index, column = 1).value = day
		report_ws_base.cell(row = index, column = 1).style = new_style
		
		set_accountID = day_accountID[day]
		size = len(set_accountID)
		line = ""
		for i in xrange(0,size):
			if 0 == i:
				line += "%s" % set_accountID[i]
			else:
				line += ";%s" % set_accountID[i]
		report_ws_base.cell(row = index, column = 2).value = line
		report_ws_base.cell(row = index, column = 2).style = new_style

		index += 1

	#记录每日操作合约
	index += 1
	report_ws_base.cell(row = index, column = 1).value = "操作合约情况："
	index += 1
	report_ws_base.cell(row = index, column = 1).value = "日期"
	report_ws_base.cell(row = index, column = 2).value = "合约"
	index += 1
	days = day_contracts.keys()
	days.sort()
	for day in days:
		report_ws_base.cell(row = index, column = 1).value = day
		report_ws_base.cell(row = index, column = 1).style = new_style
		
		set_contracts = day_contracts[day]
		size = len(set_contracts)
		line = ""
		for i in xrange(0,size):
			if 0 == i:
				line += "%s" % set_contracts[i]
			else:
				line += ";%s" % set_contracts[i]
		report_ws_base.cell(row = index, column = 2).value = line
		report_ws_base.cell(row = index, column = 2).style = new_style

		index += 1

	#记录每日交易次数
	index += 1
	report_ws_base.cell(row = index, column = 1).value = "每日交易次数："
	index += 1
	report_ws_base.cell(row = index, column = 1).value = "日期"
	report_ws_base.cell(row = index, column = 2).value = "交易次数"
	index += 1
	days = day_times.keys()
	days.sort()
	for day in days:
		report_ws_base.cell(row = index, column = 1).value = day
		report_ws_base.cell(row = index, column = 1).style = new_style
		report_ws_base.cell(row = index, column = 2).value = day_times[day]
		report_ws_base.cell(row = index, column = 2).style = new_style
		index += 1

	#在第二个sheet记录每日净值
	report_ws_profit = report_wb.create_sheet()
	report_ws_profit.title = u"每日净值"

	index_row = 1
	index_column = 1
	report_ws_profit.cell(row = index_row, column = index_column).value = "日期"
	index_column += 1
	for contract in contracts:
		report_ws_profit.cell(row = index_row, column = index_column).value = contract
		index_column += 1
	report_ws_profit.cell(row = index_row, column = index_column).value = "总净值"

	index_row += 1
	days = day_contract_profits.keys()
	days.sort()
	for day in days:
		index_column = 1
		report_ws_profit.cell(row = index_row, column = index_column).value = day

		map_profit = day_contract_profits[day]

		index_column += 1
		total_profit = 0
		for contract in contracts:
			if map_profit.has_key(contract):
				report_ws_profit.cell(row = index_row, column = index_column).value = map_profit[contract]
				total_profit += map_profit[contract]
			else:
				report_ws_profit.cell(row = index_row, column = index_column).value = 0
			index_column += 1
		report_ws_profit.cell(row = index_row, column = index_column).value = total_profit

		index_row += 1

	report_wb.save(excel_file)

def terminateProcess(processName):
    for i in WMI().Win32_Process(caption=processName):
        i.Terminate()

def do_chart(filename):
	luc = list(ascii_uppercase)
	columns = Series((luc + [i+j for i in luc for j in luc])[:256],range(1,257))
	def cellName(nRow,nCol):
	    return columns[nCol]+str(nRow)

	application = win32com.client.Dispatch('Excel.Application')
	workbook = application.Workbooks.Open(filename)
	sheets = workbook.Sheets
	sheet1 = sheets.Item(2)

	nRow = sheet1.UsedRange.Rows.Count
	nCol = sheet1.UsedRange.Columns.Count

	for x in xrange(2,nCol+1):
		chartObjectXCells = 5
		chartObjectYCells = 12
		chartObjectLeft = sheet1.Cells(2,x).Left
		chartObjectTop = sheet1.Cells(2,x).Top
		chartObjectWidth = sheet1.Cells(2,1+x+chartObjectXCells).Left - chartObjectLeft
		chartObjectHeight = sheet1.Cells(2+chartObjectYCells,1).Top - chartObjectTop
		chartObject = sheet1.ChartObjects().Add(chartObjectLeft,chartObjectTop,chartObjectWidth,chartObjectHeight)

		chart = chartObject.Chart
		chart.Legend.Position = xlLegendPositionBottom

		seriesCollection = chart.SeriesCollection()
		rangeName = cellName(2,x) + ':' + cellName(nRow,x)
		xRangeName = cellName(2,1) + ':' + cellName(nRow,1)

		series = seriesCollection.NewSeries()
		series.Name = sheet1.Cells(1,x)
		series.Values = sheet1.Range(rangeName)
		series.XValues = sheet1.Range(xRangeName)
		series.Format.Line.Weight = 1
		series.ChartType = xlLine

	workbook.Save()
	workbook.Close()

	application.Quit()
	terminateProcess("EXCEL.EXE")

def main():
	make_profit_dir()

	trade_rec_file = "1-25.txt"
	if len(sys.argv) > 1:
		trade_rec_file = sys.argv[1]

	# 读取交易记录
	records = read_records(trade_rec_file)
	if len(records) == 0:
		raise Exception('cant read trading record!')
		return

	pos_manager = positions.PositionsManager()

	# 分析交易收益
	infos = parse_records(pos_manager, records)

	# 计算每日净值
	calc_day_profit(infos["contracts"].keys(), infos["day_profit"], infos["day_contracts"], infos["day_accountID"], infos["day_times"])

	# 生成excel报告
	do_excel(infos["contracts"].keys(), infos["day_profit"], infos["day_contracts"], infos["day_accountID"], infos["day_times"])

	# 在excel中画曲线图
	homedir = os.getcwd()
	do_chart(homedir + "\\log\\test.xlsx")

	print "----------------"
	log.WriteLog("result", "合约\t平仓收益\t手续费")
	for c,v in infos["contracts"].items():
		log.WriteLog("result" ,"%s\t%.2f\t%.2f"%(c, v[0], v[1]))


if __name__ == '__main__':
	main()