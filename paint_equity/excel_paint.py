# -*- coding: utf-8 -*-

import os
import re
from openpyxl.workbook import Workbook
from openpyxl.chart import Series, Reference, LineChart
import equity_sort

# from openpyxl.drawing.drawing import Shape

import win32com.client
# from win32com.gen_py import msof,mspp,msxl
from string import ascii_uppercase
from pandas import Series

xlLegendPositionBottom = -4107
xlAreaStacked = 76         # from enum XlChartType
xlLine = 4          # from enum XlChartType

g = globals()
# for c in dir(msof.constants) : g[c] = getattr(msof.constants, c)
# for c in dir(mspp.constants) : g[c] = getattr(mspp.constants, c)
# for c in dir(win32com.client.constants) : g[c] =
# getattr(win32com.client.constants, c)

from wmi import WMI


def terminateProcess(processName):
    for i in WMI().Win32_Process(caption=processName):
        i.Terminate()

def choose_margin_files(rootdir, required_num):
    margin_part_files = []
    for parent, dirnames, filenames in os.walk(rootdir):
        # print(filenames)
        for filename in filenames:
            if filename.find("margin") != -1:
                ma = re.findall("([r ]*)(\d+\.\d+)_(\w+)\.log", filename)
                if len(ma) > 0:
                    evalute = float(ma[0][1])
                    if evalute == 0:
                        continue
                    if ma[0][0] == 'r ':
                        evalute = -evalute
                    margin_part_files.append([ma[0][2], evalute])

    margin_part_files = sorted(
        margin_part_files, key=lambda d: d[1], reverse=True)
    key_number = len(margin_part_files)

    chosen_interval = 1
    if required_num > 0:
        chosen_interval = max(1, key_number / required_num)

    margin_files = []
    for x in range(0, key_number, chosen_interval):
        new_file_name = ""
        evalute = margin_part_files[x][1]
        log_name = margin_part_files[x][0]
        if evalute >= 0:
            new_file_name = "%f_%s.log" % (evalute, log_name)
        else:
            new_file_name = "r %f_%s.log" % (-evalute, log_name)
        fullname = os.path.join(parent, new_file_name)
        margin_files.append(fullname)

    return margin_files


# def do_chart(ws, begin_line, end_line, x_column, y_column):
#   xyserie1 = Series(Reference(ws, (begin_line,y_column), (end_line,y_column)),
# labels=Reference(ws, (begin_line,x_column), (end_line,x_column)),
# color="red")

#   linechart = LineChart()
#   linechart.add_serie(xyserie1)

#   linechart.x_axis.min = 0
#   linechart.x_axis.max = 0.5
#   linechart.x_axis.unit = 0.01

#   linechart.y_axis.min = 0
#   linechart.y_axis.max = 0.5
#   linechart.y_axis.unit = 0.01

#   linechart.drawing.left = 400
#   linechart.drawing.top = 200
#   linechart.drawing.height = 400
#   linechart.drawing.width = 700

#   linechart.width = .9
#   linechart.height = .9
#   linechart.margin_top = .1
#   ws.add_chart(linechart)


def get_total_col_title(file_name):
    index = file_name.rfind(".")
    if -1 != index:
        file_name = file_name[0:index]

    title_str = ""
    list_str = file_name.split('_')
    count = len(list_str)
    for i in range(0, count):
        try:
            int(list_str[i])
        except Exception as e:
            title_str = ""
            continue

        title_str = title_str + "_" + list_str[i]

    if len(title_str) == 0:
        return "equity"

    return title_str


def do_excel(path, files_datas, align_pos, create_total):
    margin_files = files_datas.keys()
    total_fullname_xls = os.path.join(path, "total_margin.xlsx")
    if os.path.exists(total_fullname_xls):
        os.remove(total_fullname_xls)

    margin_total_wb = Workbook()
    margin_total_ws = margin_total_wb.active

    total_index = 1
    for margin_fullname in margin_files:
        col_title = get_total_col_title(margin_fullname)

        _margin_fullname_xls = margin_fullname.replace(".log", ".xlsx")
        if os.path.exists(_margin_fullname_xls):
            os.remove(_margin_fullname_xls)

        margin_wb = Workbook()
        margin_ws = margin_wb.active

        index = 1
        for _date, _equity, _margin_rate in files_datas[margin_fullname]:
            if align_pos is not None:
                _equity = _equity - files_datas[margin_fullname][align_pos][1]

            margin_ws.cell(row=index + 1, column=1).value = _date
            margin_ws.cell(row=index + 1, column=2).value = _equity
            margin_ws.cell(row=index + 1, column=3).value = _margin_rate

            if create_total:
                if 1 == total_index:
                    margin_total_ws.cell(row=index + 1, column=total_index).value = _date
                    margin_total_ws.cell(row=index + 1, column=total_index + 1).value = _equity
                else:
                    margin_total_ws.cell(row=index + 1, column=total_index + 1).value = _equity

            if 1 == index:
                margin_ws.cell(row=index, column=1).value = "date"
                margin_ws.cell(row=index, column=2).value = col_title
                margin_ws.cell(row=index, column=3).value = "margin_rate"
                if create_total:
                    margin_total_ws.cell(row=index, column=1).value = "date"
                    margin_total_ws.cell(row=index, column=total_index + 1).value = col_title

            index = index + 1

        margin_wb.save(_margin_fullname_xls)
        total_index = total_index + 1

    if create_total:
        margin_total_wb.save(total_fullname_xls)


def do_chart(rootdir):
    luc = list(ascii_uppercase)
    columns = Series(
        (luc + [i + j for i in luc for j in luc])[:256], range(1, 257))

    def cellName(nRow, nCol):
        return columns[nCol] + str(nRow)

    application = win32com.client.Dispatch('Excel.Application')
    for parent, dirnames, filenames in os.walk(rootdir):
        for filename in filenames:
            if filename.find(".xlsx") == -1:
                continue
            try:
                filename = os.path.join(parent, filename)
                workbook = application.Workbooks.Open(filename)
                sheets = workbook.Sheets
                sheet1 = sheets.Item(1)

                nRow = sheet1.UsedRange.Rows.Count
                nCol = sheet1.UsedRange.Columns.Count

                chartObjectXCells = 15
                chartObjectYCells = 50
                chartObjectLeft = sheet1.Cells(2, nCol + 2).Left
                chartObjectTop = sheet1.Cells(2, 1).Top
                chartObjectWidth = sheet1.Cells(
                    2, nCol + 2 + chartObjectXCells).Left - chartObjectLeft
                chartObjectHeight = sheet1.Cells(
                    2 + chartObjectYCells, 1).Top - chartObjectTop
                chartObject = sheet1.ChartObjects().Add(chartObjectLeft, chartObjectTop,
                                                        chartObjectWidth, chartObjectHeight)

                chart = chartObject.Chart
                chart.Legend.Position = xlLegendPositionBottom

                seriesCollection = chart.SeriesCollection()
                for i in range(2, nCol + 1):
                    rangeName = cellName(2, i) + ':' + cellName(nRow, i)
                    xRangeName = cellName(2, 1) + ':' + cellName(nRow, 1)

                    series = seriesCollection.NewSeries()
                    series.Name = sheet1.Cells(1, i)
                    series.Values = sheet1.Range(rangeName)
                    series.XValues = sheet1.Range(xRangeName)
                    if filename.find("total_margin") == -1:
                        series.Format.Line.Weight = 2
                        if 2 == i:
                            series.ChartType = xlLine
                        else:
                            series.AxisGroup = 2
                            series.ChartType = xlAreaStacked
                    else:
                        series.Format.Line.Weight = 2
                        series.ChartType = xlLine

                workbook.Save()
                workbook.Close()
            except Exception as e:
                import traceback
                print(e, traceback.print_exc())

    application.Quit()
    terminateProcess("EXCEL.EXE")


def main():
    path = "./equity"

    margin_files = choose_margin_files(path, 10)

    do_excel(path, margin_files, True)

    full_path = os.getcwd() + "\\equity"
    do_chart(full_path)


if __name__ == '__main__':
    main()
